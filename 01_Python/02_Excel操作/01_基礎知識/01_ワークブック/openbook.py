#xlsxファイルを扱うライブラリを読み込む
import openpyxl

import os

#カレントディレクトリの変更
os.chdir('C:\\Users\\empla\\OneDrive\\デスクトップ\\work\\01_Python\\02_Excel操作\\01_基礎知識\\01_ワークブック')

try:

    wb = openpyxl.load_workbook("..\99_Excel\File_Dialog.xlsx")

except:

    wb = openpyxl.Workbook()

finally:

    print(wb.sheetnames)