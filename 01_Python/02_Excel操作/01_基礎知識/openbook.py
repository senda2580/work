#xlsxファイルを扱うライブラリを読み込む
import openpyxl

try:

    wb = openpyxl.load_workbook("Excel\File_Dialog.xlsx")

except:

    wb = openpyxl.Workbook()

finally:

    print(wb.sheetnames)